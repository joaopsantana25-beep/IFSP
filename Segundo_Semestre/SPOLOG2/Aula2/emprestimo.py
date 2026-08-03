import json
import os
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from livro import (
    buscar_livro_por_codigo,
    listar_livros,
    atualizar_disponibilidade_livro
)

from usuario import (
    buscar_usuario_por_codigo,
    listar_usuarios
)


PASTA_DADOS = "dados"
PASTA_PLANILHAS = "planilhas"
ARQUIVO_EMPRESTIMOS = os.path.join(PASTA_DADOS, "emprestimos.json")
PLANILHA_EMPRESTIMOS = os.path.join(
    PASTA_PLANILHAS,
    "cadastro_emprestimos.xlsx"
)


class Emprestimo:
    """
    Representa um emprestimo realizado pela biblioteca.
    """

    def __init__(
        self,
        codigo,
        codigo_livro,
        titulo_livro,
        codigo_usuario,
        nome_usuario,
        data_emprestimo,
        data_prevista_devolucao,
        situacao="Ativo"
    ):
        self.codigo = codigo
        self.codigo_livro = codigo_livro
        self.titulo_livro = titulo_livro
        self.codigo_usuario = codigo_usuario
        self.nome_usuario = nome_usuario
        self.data_emprestimo = data_emprestimo
        self.data_prevista_devolucao = data_prevista_devolucao
        self.situacao = situacao

    def para_dicionario(self):
        """
        Converte o objeto Emprestimo em um dicionario.
        """
        return {
            "codigo": self.codigo,
            "codigo_livro": self.codigo_livro,
            "titulo_livro": self.titulo_livro,
            "codigo_usuario": self.codigo_usuario,
            "nome_usuario": self.nome_usuario,
            "data_emprestimo": self.data_emprestimo,
            "data_prevista_devolucao": self.data_prevista_devolucao,
            "situacao": self.situacao
        }


def preparar_armazenamento_emprestimos():
    """
    Cria a pasta dados e o arquivo emprestimos.json.
    """
    os.makedirs(PASTA_DADOS, exist_ok=True)

    if not os.path.exists(ARQUIVO_EMPRESTIMOS):
        with open(ARQUIVO_EMPRESTIMOS, "w", encoding="utf-8") as arquivo:
            json.dump([], arquivo, ensure_ascii=False, indent=4)


def carregar_emprestimos():
    """
    Le os emprestimos gravados no arquivo JSON.
    """
    preparar_armazenamento_emprestimos()

    try:
        with open(ARQUIVO_EMPRESTIMOS, "r", encoding="utf-8") as arquivo:
            return json.load(arquivo)
    except (json.JSONDecodeError, FileNotFoundError):
        return []


def salvar_emprestimos(emprestimos):
    """
    Salva os emprestimos no arquivo JSON.
    """
    preparar_armazenamento_emprestimos()

    with open(ARQUIVO_EMPRESTIMOS, "w", encoding="utf-8") as arquivo:
        json.dump(emprestimos, arquivo, ensure_ascii=False, indent=4)


def gerar_proximo_codigo_emprestimo():
    """
    Gera o proximo codigo de emprestimo.
    """
    emprestimos = carregar_emprestimos()

    if not emprestimos:
        return 1

    maior_codigo = max(
        emprestimo["codigo"]
        for emprestimo in emprestimos
    )

    return maior_codigo + 1


def solicitar_codigo(mensagem):
    """
    Solicita um codigo numerico e trata entradas invalidas.
    """
    while True:
        valor = input(mensagem).strip()

        if not valor.isdigit():
            print("Informe apenas numeros.")
            continue

        return int(valor)


def realizar_emprestimo():
    """
    Realiza o emprestimo de um livro para um usuario.
    """
    print("\n" + "=" * 55)
    print("REALIZACAO DE EMPRESTIMO")
    print("=" * 55)

    listar_livros()

    codigo_livro = solicitar_codigo(
        "\nInforme o codigo do livro: "
    )

    livro = buscar_livro_por_codigo(codigo_livro)

    if livro is None:
        print("\nLivro nao encontrado.")
        return

    if not livro["disponivel"]:
        print("\nEsse livro ja esta emprestado.")
        return

    listar_usuarios()

    codigo_usuario = solicitar_codigo(
        "\nInforme o codigo do usuario: "
    )

    usuario = buscar_usuario_por_codigo(codigo_usuario)

    if usuario is None:
        print("\nUsuario nao encontrado.")
        return

    data_atual = datetime.now()
    data_prevista = data_atual + timedelta(days=7)

    novo_emprestimo = Emprestimo(
        codigo=gerar_proximo_codigo_emprestimo(),
        codigo_livro=livro["codigo"],
        titulo_livro=livro["titulo"],
        codigo_usuario=usuario["codigo"],
        nome_usuario=usuario["nome"],
        data_emprestimo=data_atual.strftime("%d/%m/%Y"),
        data_prevista_devolucao=data_prevista.strftime("%d/%m/%Y"),
        situacao="Ativo"
    )

    emprestimos = carregar_emprestimos()
    emprestimos.append(novo_emprestimo.para_dicionario())
    salvar_emprestimos(emprestimos)

    atualizar_disponibilidade_livro(
        codigo=livro["codigo"],
        disponivel=False
    )

    print("\nEmprestimo realizado com sucesso.")
    print(f"Livro: {livro['titulo']}")
    print(f"Usuario: {usuario['nome']}")
    print(
        "Data prevista para devolucao: "
        f"{novo_emprestimo.data_prevista_devolucao}"
    )


def exportar_emprestimos_excel():
    """
    Exporta os emprestimos cadastrados para uma planilha Excel.
    """
    emprestimos = carregar_emprestimos()

    if not emprestimos:
        print("\nNao existem emprestimos para exportar.")
        return

    os.makedirs(PASTA_PLANILHAS, exist_ok=True)

    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Emprestimos"

    cabecalhos = [
        "Codigo do emprestimo",
        "Codigo do livro",
        "Livro",
        "Codigo do usuario",
        "Usuario",
        "Data do emprestimo",
        "Devolucao prevista",
        "Situacao"
    ]

    planilha.append(cabecalhos)

    preenchimento = PatternFill(
        fill_type="solid",
        fgColor="548235"
    )

    fonte = Font(
        color="FFFFFF",
        bold=True
    )

    for celula in planilha[1]:
        celula.fill = preenchimento
        celula.font = fonte
        celula.alignment = Alignment(horizontal="center")

    for emprestimo in emprestimos:
        planilha.append([
            emprestimo["codigo"],
            emprestimo["codigo_livro"],
            emprestimo["titulo_livro"],
            emprestimo["codigo_usuario"],
            emprestimo["nome_usuario"],
            emprestimo["data_emprestimo"],
            emprestimo["data_prevista_devolucao"],
            emprestimo["situacao"]
        ])

    larguras = {
        "A": 22,
        "B": 18,
        "C": 35,
        "D": 20,
        "E": 30,
        "F": 20,
        "G": 22,
        "H": 15
    }

    for coluna, largura in larguras.items():
        planilha.column_dimensions[coluna].width = largura

    planilha.freeze_panes = "A2"
    planilha.auto_filter.ref = planilha.dimensions

    workbook.save(PLANILHA_EMPRESTIMOS)

    print("\nCadastro de emprestimos exportado com sucesso.")
    print(f"Arquivo criado: {PLANILHA_EMPRESTIMOS}")

