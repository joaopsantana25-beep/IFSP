import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


PASTA_DADOS = "dados"
PASTA_PLANILHAS = "planilhas"
ARQUIVO_LIVROS = os.path.join(PASTA_DADOS, "livros.json")
PLANILHA_LIVROS = os.path.join(PASTA_PLANILHAS, "cadastro_livros.xlsx")


class Livro:
    """
    Representa um livro cadastrado na biblioteca.
    """

    def __init__(self, codigo, titulo, autor, editora, ano, disponivel=True):
        self.codigo = codigo
        self.titulo = titulo
        self.autor = autor
        self.editora = editora
        self.ano = ano
        self.disponivel = disponivel

    def para_dicionario(self):
        """
        Converte o objeto Livro para um dicionario.
        Isso permite salvar os dados em um arquivo JSON.
        """
        return {
            "codigo": self.codigo,
            "titulo": self.titulo,
            "autor": self.autor,
            "editora": self.editora,
            "ano": self.ano,
            "disponivel": self.disponivel
        }


def preparar_armazenamento_livros():
    """
    Cria a pasta de dados e o arquivo livros.json,
    caso ainda nao existam.
    """
    os.makedirs(PASTA_DADOS, exist_ok=True)

    if not os.path.exists(ARQUIVO_LIVROS):
        with open(ARQUIVO_LIVROS, "w", encoding="utf-8") as arquivo:
            json.dump([], arquivo, ensure_ascii=False, indent=4)


def carregar_livros():
    """
    Le os livros armazenados no arquivo JSON.
    Retorna uma lista de dicionarios.
    """
    preparar_armazenamento_livros()

    try:
        with open(ARQUIVO_LIVROS, "r", encoding="utf-8") as arquivo:
            return json.load(arquivo)
    except (json.JSONDecodeError, FileNotFoundError):
        return []


def salvar_livros(livros):
    """
    Salva a lista de livros no arquivo JSON.
    """
    preparar_armazenamento_livros()

    with open(ARQUIVO_LIVROS, "w", encoding="utf-8") as arquivo:
        json.dump(livros, arquivo, ensure_ascii=False, indent=4)


def gerar_proximo_codigo():
    """
    Gera automaticamente o proximo codigo de livro.
    """
    livros = carregar_livros()

    if not livros:
        return 1

    maior_codigo = max(livro["codigo"] for livro in livros)
    return maior_codigo + 1


def solicitar_ano():
    """
    Solicita e valida o ano de publicacao do livro.
    """
    while True:
        ano = input("Ano de publicacao: ").strip()

        if not ano.isdigit():
            print("O ano deve conter apenas numeros.")
            continue

        ano_numerico = int(ano)

        if ano_numerico <= 0:
            print("Informe um ano valido.")
            continue

        return ano_numerico


def cadastrar_livro():
    """
    Solicita os dados do livro e realiza o cadastro.
    """
    print("\n" + "=" * 50)
    print("CADASTRO DE LIVRO")
    print("=" * 50)

    titulo = input("Titulo do livro: ").strip()
    autor = input("Autor: ").strip()
    editora = input("Editora: ").strip()
    ano = solicitar_ano()

    if not titulo or not autor or not editora:
        print("\nTitulo, autor e editora sao obrigatorios.")
        return

    livros = carregar_livros()
    codigo = gerar_proximo_codigo()

    novo_livro = Livro(
        codigo=codigo,
        titulo=titulo,
        autor=autor,
        editora=editora,
        ano=ano,
        disponivel=True
    )

    livros.append(novo_livro.para_dicionario())
    salvar_livros(livros)

    print("\nLivro cadastrado com sucesso.")
    print(f"Codigo gerado: {codigo}")


def listar_livros():
    """
    Exibe todos os livros cadastrados.
    """
    livros = carregar_livros()

    print("\n" + "=" * 90)
    print("LIVROS CADASTRADOS")
    print("=" * 90)

    if not livros:
        print("Nenhum livro cadastrado.")
        return

    print(
        f"{'CODIGO':<8}"
        f"{'TITULO':<28}"
        f"{'AUTOR':<22}"
        f"{'ANO':<8}"
        f"{'SITUACAO':<15}"
    )

    print("-" * 90)

    for livro in livros:
        situacao = "Disponivel" if livro["disponivel"] else "Emprestado"

        print(
            f"{livro['codigo']:<8}"
            f"{livro['titulo'][:26]:<28}"
            f"{livro['autor'][:20]:<22}"
            f"{livro['ano']:<8}"
            f"{situacao:<15}"
        )


def buscar_livro_por_codigo(codigo):
    """
    Procura um livro usando seu codigo.
    Retorna o livro encontrado ou None.
    """
    livros = carregar_livros()

    for livro in livros:
        if livro["codigo"] == codigo:
            return livro

    return None


def atualizar_disponibilidade_livro(codigo, disponivel):
    """
    Atualiza a situacao de disponibilidade de um livro.
    """
    livros = carregar_livros()

    for livro in livros:
        if livro["codigo"] == codigo:
            livro["disponivel"] = disponivel
            salvar_livros(livros)
            return True

    return False


def exportar_livros_excel():
    """
    Exporta os livros cadastrados para uma planilha Excel.
    """
    livros = carregar_livros()

    if not livros:
        print("\nNao existem livros para exportar.")
        return

    os.makedirs(PASTA_PLANILHAS, exist_ok=True)

    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Livros"

    cabecalhos = [
        "Codigo",
        "Titulo",
        "Autor",
        "Editora",
        "Ano",
        "Situacao"
    ]

    planilha.append(cabecalhos)

    preenchimento = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    fonte = Font(
        color="FFFFFF",
        bold=True
    )

    for celula in planilha[1]:
        celula.fill = preenchimento
        celula.font = fonte
        celula.alignment = Alignment(horizontal="center")

    for livro in livros:
        situacao = "Disponivel" if livro["disponivel"] else "Emprestado"

        planilha.append([
            livro["codigo"],
            livro["titulo"],
            livro["autor"],
            livro["editora"],
            livro["ano"],
            situacao
        ])

    larguras = {
        "A": 12,
        "B": 35,
        "C": 30,
        "D": 25,
        "E": 12,
        "F": 18
    }

    for coluna, largura in larguras.items():
        planilha.column_dimensions[coluna].width = largura

    planilha.freeze_panes = "A2"
    planilha.auto_filter.ref = planilha.dimensions

    workbook.save(PLANILHA_LIVROS)

    print("\nCadastro de livros exportado com sucesso.")
    print(f"Arquivo criado: {PLANILHA_LIVROS}")